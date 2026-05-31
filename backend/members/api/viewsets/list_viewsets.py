"""
ViewSet for MemberList and MemberListEntry management.
"""

from datetime import date
from io import BytesIO

import openpyxl
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import OpenApiParameter, extend_schema, extend_schema_view
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.renderers import BaseRenderer
from rest_framework.response import Response

from jf_manager_backend.permissions import DepartmentRoleModelPermissions
from members.api.serializers.list_serializers import (
    CreateFromEventTypeInputSerializer,
    MemberListCreateUpdateSerializer,
    MemberListDetailSerializer,
    MemberListSerializer,
)
from members.api.viewsets.member_viewsets import MEMBER_EXPORT_COLUMNS, MEMBER_EXPORT_DEFAULT_COLUMNS
from members.api_serializers import AttachmentSerializer
from members.models import Attachment, Event, Member, MemberList, MemberListEntry


class PassthroughRenderer(BaseRenderer):
    """Return data as-is for binary responses."""

    media_type = "*/*"
    format = "binary"

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


LIST_EXTRA_COLUMNS = {
    "list_checked": "Anwesend",
    "list_checked_at": "Anwesend seit",
    "list_notes": "Notiz (Liste)",
}
LIST_EXTRA_DEFAULT_COLUMNS = ["list_checked", "list_notes"]

ALL_LIST_EXPORT_COLUMNS = {**MEMBER_EXPORT_COLUMNS, **LIST_EXTRA_COLUMNS}


@extend_schema_view(
    list=extend_schema(summary="List all member lists"),
    retrieve=extend_schema(summary="Get a member list with all entries"),
    create=extend_schema(summary="Create a new member list"),
    update=extend_schema(summary="Update a member list"),
    partial_update=extend_schema(summary="Partially update a member list"),
    destroy=extend_schema(summary="Delete a member list"),
)
class MemberListViewSet(viewsets.ModelViewSet):
    queryset = MemberList.objects.all()
    permission_classes = [IsAuthenticated, DepartmentRoleModelPermissions]

    def get_serializer_class(self):
        if self.action in ("create", "update", "partial_update"):
            return MemberListCreateUpdateSerializer
        if self.action == "retrieve":
            return MemberListDetailSerializer
        return MemberListSerializer

    @action(detail=True, methods=["get", "post"], url_path="attachments", parser_classes=[MultiPartParser, FormParser])
    def attachments(self, request, pk=None):
        """
        GET: List attachments for a member list
        POST: Upload an attachment for a member list
        """
        member_list = self.get_object()

        if request.method == "GET":
            attachments = member_list.attachments.all().order_by("-uploaded_at")
            serializer = AttachmentSerializer(attachments, many=True, context={"request": request})
            return Response(serializer.data)

        if not request.user.has_perm("members.change_memberlist"):
            return Response({"detail": "Keine Berechtigung zum Bearbeiten."}, status=status.HTTP_403_FORBIDDEN)

        file_obj = request.FILES.get("file")
        name = request.data.get("name") or (file_obj.name if file_obj else None)
        description = request.data.get("description", "")

        if not file_obj or not name:
            return Response({"detail": "Datei und Name sind erforderlich."}, status=status.HTTP_400_BAD_REQUEST)

        attachment = Attachment.objects.create(
            content_object=member_list,
            file=file_obj,
            name=name,
            description=description,
            uploaded_by=request.user,
        )

        serializer = AttachmentSerializer(attachment, context={"request": request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["delete"], url_path="attachments/(?P<attachment_id>[^/.]+)")
    def delete_attachment(self, request, pk=None, attachment_id=None):
        """Delete an attachment from a member list."""
        member_list = self.get_object()

        if not request.user.has_perm("members.change_memberlist"):
            return Response({"detail": "Keine Berechtigung zum Löschen."}, status=status.HTTP_403_FORBIDDEN)

        attachment = get_object_or_404(member_list.attachments, pk=attachment_id)
        attachment.delete()

        return Response(status=status.HTTP_204_NO_CONTENT)

    # ── Membership management ──────────────────────────────────────────────

    @action(detail=True, methods=["post"])
    def add_member(self, request, pk=None):
        """POST /api/v1/member-lists/{id}/add_member/ — add a member to this list."""
        member_list = self.get_object()
        member_id = request.data.get("member_id")
        if not member_id:
            return Response({"error": "member_id erforderlich."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            member = Member.objects.get(pk=member_id)
        except Member.DoesNotExist:
            return Response({"error": "Mitglied nicht gefunden."}, status=status.HTTP_404_NOT_FOUND)

        _, created = MemberListEntry.objects.get_or_create(
            member_list=member_list,
            member=member,
        )
        return Response(
            {"added": created, "member_count": member_list.member_count},
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"])
    def remove_member(self, request, pk=None):
        """POST /api/v1/member-lists/{id}/remove_member/ — remove a member from this list."""
        member_list = self.get_object()
        member_id = request.data.get("member_id")
        deleted, _ = MemberListEntry.objects.filter(member_list=member_list, member_id=member_id).delete()
        return Response({"removed": bool(deleted), "member_count": member_list.member_count})

    @action(detail=True, methods=["post"])
    def bulk_add(self, request, pk=None):
        """POST /api/v1/member-lists/{id}/bulk_add/ — add multiple members at once."""
        member_list = self.get_object()
        member_ids = request.data.get("member_ids", [])
        if not isinstance(member_ids, list):
            return Response({"error": "member_ids muss eine Liste sein."}, status=status.HTTP_400_BAD_REQUEST)

        added = 0
        for member_id in member_ids:
            try:
                member = Member.objects.get(pk=member_id)
                _, created = MemberListEntry.objects.get_or_create(member_list=member_list, member=member)
                if created:
                    added += 1
            except Member.DoesNotExist:
                pass

        return Response({"added": added, "member_count": member_list.member_count})

    # ── Check/uncheck ──────────────────────────────────────────────────────

    @action(detail=True, methods=["post"])
    def toggle_check(self, request, pk=None):
        """POST /api/v1/member-lists/{id}/toggle_check/ — toggle checked state for one member."""
        member_list = self.get_object()
        member_id = request.data.get("member_id")
        try:
            entry = MemberListEntry.objects.get(member_list=member_list, member_id=member_id)
        except MemberListEntry.DoesNotExist:
            return Response({"error": "Mitglied nicht in dieser Liste."}, status=status.HTTP_404_NOT_FOUND)
        entry.toggle_check()
        return Response({"checked": entry.checked, "checked_at": entry.checked_at})

    @action(detail=True, methods=["post"])
    def set_check(self, request, pk=None):
        """POST /api/v1/member-lists/{id}/set_check/ — set checked state for one member explicitly."""
        member_list = self.get_object()
        member_id = request.data.get("member_id")
        checked = request.data.get("checked")
        if checked is None:
            return Response({"error": "checked erforderlich."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            entry = MemberListEntry.objects.get(member_list=member_list, member_id=member_id)
        except MemberListEntry.DoesNotExist:
            return Response({"error": "Mitglied nicht in dieser Liste."}, status=status.HTTP_404_NOT_FOUND)
        entry.checked = bool(checked)
        entry.checked_at = timezone.now() if entry.checked else None
        entry.save(update_fields=["checked", "checked_at"])
        return Response({"checked": entry.checked, "checked_at": entry.checked_at})

    @action(detail=True, methods=["post"])
    def check_all(self, request, pk=None):
        """POST /api/v1/member-lists/{id}/check_all/ — mark all members as checked."""
        member_list = self.get_object()
        now = timezone.now()
        member_list.entries.update(checked=True, checked_at=now)
        return Response({"checked_count": member_list.member_count})

    @action(detail=True, methods=["post"])
    def uncheck_all(self, request, pk=None):
        """POST /api/v1/member-lists/{id}/uncheck_all/ — reset all check states."""
        member_list = self.get_object()
        member_list.entries.update(checked=False, checked_at=None)
        return Response({"checked_count": 0})

    @action(detail=True, methods=["patch"])
    def update_entry_notes(self, request, pk=None):
        """PATCH /api/v1/member-lists/{id}/update_entry_notes/ — set notes on a list entry."""
        member_list = self.get_object()
        member_id = request.data.get("member_id")
        notes = request.data.get("notes", "")
        updated = MemberListEntry.objects.filter(member_list=member_list, member_id=member_id).update(notes=notes)
        if not updated:
            return Response({"error": "Eintrag nicht gefunden."}, status=status.HTTP_404_NOT_FOUND)
        return Response({"notes": notes})

    @action(detail=False, methods=["post"], url_path="create_from_event_type")
    def create_from_event_type(self, request):
        """
        POST /api/v1/member-lists/create_from_event_type/

        Create a MemberList pre-populated by filtering members based on whether
        they have (or don't have) a log entry of a given type, optionally within
        a date range.
        """
        serializer = CreateFromEventTypeInputSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        name = data["name"]
        description = data.get("description", "")
        event_type_id = data.get("event_type_id")
        invert = data.get("invert", False)
        date_from = data.get("date_from")
        date_to = data.get("date_to")

        user = request.user
        is_org_wide = user.is_staff or user.is_superuser or user.has_perm("departments.can_access_all_departments")

        # Determine member base set scoped to the user's accessible departments
        dept_raw = request.query_params.get("department")
        if is_org_wide:
            if dept_raw:
                try:
                    dept_id = int(dept_raw)
                    base_members = Member.objects.filter(departments__id=dept_id)
                except (ValueError, TypeError):
                    base_members = Member.objects.all()
            else:
                base_members = Member.objects.all()
        else:
            dept_ids = list(user.department_roles.values_list("department_id", flat=True))
            if dept_raw:
                try:
                    dept_id = int(dept_raw)
                    if dept_id not in dept_ids:
                        return Response(
                            {"detail": "Keine Berechtigung für diese Abteilung."},
                            status=status.HTTP_403_FORBIDDEN,
                        )
                    base_members = Member.objects.filter(departments__id=dept_id)
                except (ValueError, TypeError):
                    base_members = Member.objects.filter(departments__id__in=dept_ids)
            else:
                base_members = Member.objects.filter(departments__id__in=dept_ids)

        base_members = base_members.distinct()

        # Filter events matching the criteria
        events_qs = Event.objects.filter(member__in=base_members)
        if event_type_id is not None:
            events_qs = events_qs.filter(type_id=event_type_id)
        if date_from:
            events_qs = events_qs.filter(datetime__gte=date_from)
        if date_to:
            events_qs = events_qs.filter(datetime__lte=date_to)

        member_ids_with_events = set(events_qs.values_list("member_id", flat=True))

        if invert:
            selected_members = list(base_members.exclude(id__in=member_ids_with_events))
        else:
            selected_members = list(base_members.filter(id__in=member_ids_with_events))

        member_list = MemberList.objects.create(name=name, description=description)
        MemberListEntry.objects.bulk_create(
            [MemberListEntry(member_list=member_list, member=m) for m in selected_members]
        )

        output_serializer = MemberListSerializer(member_list)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    @extend_schema(
        summary="Export member list entries to Excel with column selection",
        parameters=[
            OpenApiParameter(
                "columns",
                OpenApiTypes.STR,
                description=(
                    "Comma-separated list of column keys to include. "
                    f"Available: {', '.join(ALL_LIST_EXPORT_COLUMNS.keys())}. "
                    "Defaults to basic member fields + checked/notes when omitted."
                ),
            ),
        ],
    )
    @action(detail=True, methods=["get"], url_path="export-excel", renderer_classes=[PassthroughRenderer])
    def export_excel(self, request, pk=None):
        if not request.user.has_perm("members.view_memberlist"):
            return Response({"error": "Keine Berechtigung für Listen-Export"}, status=403)

        member_list = self.get_object()

        # Determine which columns to export
        columns_param = request.query_params.get("columns", "")
        if columns_param:
            requested = [c.strip() for c in columns_param.split(",") if c.strip()]
            selected_columns = [c for c in requested if c in ALL_LIST_EXPORT_COLUMNS]
        else:
            selected_columns = [*MEMBER_EXPORT_DEFAULT_COLUMNS, *LIST_EXTRA_DEFAULT_COLUMNS]
        if not selected_columns:
            selected_columns = [*MEMBER_EXPORT_DEFAULT_COLUMNS, *LIST_EXTRA_DEFAULT_COLUMNS]

        entries = (
            MemberListEntry.objects.filter(member_list=member_list)
            .select_related("member", "member__status", "member__group")
            .prefetch_related("member__parent_set", "member__departments")
            .order_by("member__lastname", "member__name")
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = member_list.name[:31]  # Excel sheet name limit

        header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, col_key in enumerate(selected_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=ALL_LIST_EXPORT_COLUMNS[col_key])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        today = date.today()

        for row_idx, entry in enumerate(entries, start=2):
            member = entry.member
            parents = list(member.parent_set.all()[:2])
            p1 = parents[0] if len(parents) > 0 else None
            p2 = parents[1] if len(parents) > 1 else None

            for col_idx, col_key in enumerate(selected_columns, start=1):
                value: str | int = ""
                if col_key == "list_checked":
                    value = "Ja" if entry.checked else "Nein"
                elif col_key == "list_checked_at":
                    value = entry.checked_at.astimezone().strftime("%d.%m.%Y %H:%M") if entry.checked_at else ""
                elif col_key == "list_notes":
                    value = entry.notes or ""
                elif col_key == "name":
                    value = member.name
                elif col_key == "lastname":
                    value = member.lastname
                elif col_key == "gender":
                    value = {"male": "Männlich", "female": "Weiblich", "diverse": "Divers"}.get(member.gender, "")
                elif col_key == "birthday":
                    value = member.birthday.strftime("%d.%m.%Y") if member.birthday else ""
                elif col_key == "age":
                    if member.birthday:
                        value = (
                            today.year
                            - member.birthday.year
                            - ((today.month, today.day) < (member.birthday.month, member.birthday.day))
                        )
                elif col_key == "email":
                    value = member.email
                elif col_key == "phone":
                    value = member.phone
                elif col_key == "mobile":
                    value = member.mobile
                elif col_key == "street":
                    value = member.street
                elif col_key == "zip_code":
                    value = member.zip_code
                elif col_key == "city":
                    value = member.city
                elif col_key == "joined":
                    value = member.joined.strftime("%d.%m.%Y") if member.joined else ""
                elif col_key == "status":
                    value = member.status.name if member.status else ""
                elif col_key == "group":
                    value = member.group.name if member.group else ""
                elif col_key == "identityCardNumber":
                    value = member.identityCardNumber
                elif col_key == "canSwimm":
                    value = "Ja" if member.canSwimm else "Nein"
                elif col_key == "notes":
                    value = member.notes
                elif col_key == "departments":
                    value = ", ".join(d.name for d in member.departments.all())
                elif col_key == "parent1_name":
                    value = p1.name if p1 else ""
                elif col_key == "parent1_lastname":
                    value = p1.lastname if p1 else ""
                elif col_key == "parent1_email":
                    value = p1.email if p1 else ""
                elif col_key == "parent1_email2":
                    value = p1.email2 if p1 else ""
                elif col_key == "parent1_phone":
                    value = p1.phone if p1 else ""
                elif col_key == "parent1_mobile":
                    value = p1.mobile if p1 else ""
                elif col_key == "parent1_street":
                    value = p1.street if p1 else ""
                elif col_key == "parent1_zip_code":
                    value = p1.zip_code if p1 else ""
                elif col_key == "parent1_city":
                    value = p1.city if p1 else ""
                elif col_key == "parent2_name":
                    value = p2.name if p2 else ""
                elif col_key == "parent2_lastname":
                    value = p2.lastname if p2 else ""
                elif col_key == "parent2_email":
                    value = p2.email if p2 else ""
                elif col_key == "parent2_email2":
                    value = p2.email2 if p2 else ""
                elif col_key == "parent2_phone":
                    value = p2.phone if p2 else ""
                elif col_key == "parent2_mobile":
                    value = p2.mobile if p2 else ""
                elif col_key == "parent2_street":
                    value = p2.street if p2 else ""
                elif col_key == "parent2_zip_code":
                    value = p2.zip_code if p2 else ""
                elif col_key == "parent2_city":
                    value = p2.city if p2 else ""

                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (capped at 50)
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        date_str = today.strftime("%Y-%m-%d")
        safe_name = member_list.name.replace("/", "-").replace("\\", "-")[:40]
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{safe_name}_{date_str}.xlsx"'
        return response
