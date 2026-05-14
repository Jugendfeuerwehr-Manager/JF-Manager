"""
MemberViewSet — CRUD + custom statistics/export actions.
"""

from datetime import date
from io import BytesIO

import openpyxl
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import OpenApiParameter, extend_schema, extend_schema_view
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import filters, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.renderers import BaseRenderer
from rest_framework.response import Response

from departments.mixins import DepartmentScopeViewSetMixin
from jf_manager_backend.permissions import DepartmentRoleModelPermissions
from members.api_serializers import (
    AttachmentSerializer,
    EventSerializer,
    MemberCreateUpdateSerializer,
    MemberDetailSerializer,
    MemberListSerializer,
    ParentSerializer,
)
from members.models import Attachment, Event, Group, Member, Status

MEMBER_EXPORT_COLUMNS = {
    "name": "Vorname",
    "lastname": "Nachname",
    "gender": "Geschlecht",
    "birthday": "Geburtsdatum",
    "age": "Alter",
    "email": "E-Mail",
    "phone": "Telefon",
    "mobile": "Mobil",
    "street": "Straße",
    "zip_code": "PLZ",
    "city": "Stadt",
    "joined": "Eingetreten am",
    "status": "Status",
    "group": "Gruppe",
    "identityCardNumber": "Ausweisnummer",
    "canSwimm": "Schwimmer",
    "notes": "Notizen",
    "departments": "Abteilungen",
    "parent1_name": "Elternteil 1 Vorname",
    "parent1_lastname": "Elternteil 1 Nachname",
    "parent1_email": "Elternteil 1 E-Mail",
    "parent1_email2": "Elternteil 1 E-Mail 2",
    "parent1_phone": "Elternteil 1 Telefon",
    "parent1_mobile": "Elternteil 1 Mobil",
    "parent1_street": "Elternteil 1 Straße",
    "parent1_zip_code": "Elternteil 1 PLZ",
    "parent1_city": "Elternteil 1 Stadt",
    "parent2_name": "Elternteil 2 Vorname",
    "parent2_lastname": "Elternteil 2 Nachname",
    "parent2_email": "Elternteil 2 E-Mail",
    "parent2_email2": "Elternteil 2 E-Mail 2",
    "parent2_phone": "Elternteil 2 Telefon",
    "parent2_mobile": "Elternteil 2 Mobil",
    "parent2_street": "Elternteil 2 Straße",
    "parent2_zip_code": "Elternteil 2 PLZ",
    "parent2_city": "Elternteil 2 Stadt",
}

MEMBER_EXPORT_DEFAULT_COLUMNS = [
    "name",
    "lastname",
    "gender",
    "birthday",
    "email",
    "phone",
    "mobile",
    "street",
    "zip_code",
    "city",
    "joined",
    "status",
    "group",
]


class PassthroughRenderer(BaseRenderer):
    """Return data as-is for binary responses."""

    media_type = "*/*"
    format = "binary"

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


@extend_schema_view(
    list=extend_schema(
        summary="List all members",
        description="Get a paginated list of all members with filtering and sorting",
        parameters=[
            OpenApiParameter("search", OpenApiTypes.STR, description="Search by name, lastname, email"),
            OpenApiParameter("status", OpenApiTypes.INT, description="Filter by status ID"),
            OpenApiParameter("group", OpenApiTypes.INT, description="Filter by group ID"),
            OpenApiParameter("ordering", OpenApiTypes.STR, description="Order by: name, lastname, birthday, joined"),
        ],
    ),
    retrieve=extend_schema(summary="Get member details"),
    create=extend_schema(summary="Create new member"),
    update=extend_schema(summary="Update member"),
    partial_update=extend_schema(summary="Partially update member"),
    destroy=extend_schema(summary="Delete member"),
)
class MemberViewSet(DepartmentScopeViewSetMixin, viewsets.ModelViewSet):
    queryset = Member.objects.select_related("status", "group", "storage_location").prefetch_related(
        "parent_set", "departments"
    )
    permission_classes = [IsAuthenticated, DepartmentRoleModelPermissions]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["status", "group", "canSwimm", "gender"]
    search_fields = ["name", "lastname", "email", "identityCardNumber"]
    ordering_fields = ["name", "lastname", "birthday", "joined"]
    ordering = ["lastname", "name"]
    # Members use M2M so we override get_queryset below
    department_field = "departments"

    def get_queryset(self):
        """Filter members by department using M2M lookup."""
        user = self.request.user
        base_qs = Member.objects.select_related("status", "group", "storage_location").prefetch_related(
            "parent_set", "departments"
        )

        if self._user_is_org_wide(user):
            requested_dept = self._resolve_requested_department(user)
            if requested_dept is not None:
                return base_qs.filter(departments__id=requested_dept).distinct()
            return base_qs

        allowed_ids = self._user_department_ids(user)
        requested_dept = self._resolve_requested_department(user)

        if requested_dept is not None:
            return base_qs.filter(departments__id=requested_dept).distinct()

        # Department-scoped users only see members in their departments.
        # Members with no department are NOT surfaced (include_central_records=False).
        return base_qs.filter(departments__id__in=allowed_ids).distinct()

    def perform_create(self, serializer):
        """Auto-assign department on create for dept-scoped users."""
        instance = serializer.save()
        user = self.request.user
        requested_dept = self._resolve_requested_department(user)
        if requested_dept is not None:
            from departments.models import Department

            try:
                dept = Department.objects.get(pk=requested_dept)
                instance.departments.add(dept)
            except Department.DoesNotExist:
                pass
        elif not self._user_is_org_wide(user):
            allowed_ids = self._user_department_ids(user)
            if len(allowed_ids) == 1:
                from departments.models import Department

                try:
                    dept = Department.objects.get(pk=allowed_ids[0])
                    instance.departments.add(dept)
                except Department.DoesNotExist:
                    pass

    def get_serializer_class(self):
        if self.action == "list":
            return MemberListSerializer
        elif self.action in ["create", "update", "partial_update"]:
            return MemberCreateUpdateSerializer
        return MemberDetailSerializer

    @extend_schema(
        summary="Get member statistics",
        description="Statistics about members: total, by gender, status, group, age, swim capability",
    )
    @action(detail=False, methods=["get"])
    def statistics(self, request):
        from collections import Counter
        from datetime import date

        from django.db.models import Count

        qs = self.queryset
        total = qs.count()

        # Gender distribution
        gender_counts = {g: 0 for g in ["male", "female", "diverse", ""]}
        for row in qs.values("gender").annotate(count=Count("id")):
            gender_counts[row["gender"] or ""] = row["count"]

        # Status distribution
        by_status = []
        for status_obj in Status.objects.all():
            count = qs.filter(status=status_obj).count()
            by_status.append({"name": status_obj.name, "color": status_obj.color, "count": count})
        no_status_count = qs.filter(status__isnull=True).count()
        if no_status_count:
            by_status.append({"name": "Kein Status", "color": "#aaaaaa", "count": no_status_count})

        # Group distribution
        by_group = []
        for group_obj in Group.objects.all():
            count = qs.filter(group=group_obj).count()
            by_group.append({"name": group_obj.name, "count": count})
        no_group_count = qs.filter(group__isnull=True).count()
        if no_group_count:
            by_group.append({"name": "Keine Gruppe", "count": no_group_count})

        # Age statistics
        today = date.today()
        ages = []
        for birthday in qs.filter(birthday__isnull=False).values_list("birthday", flat=True):
            age = today.year - birthday.year - ((today.month, today.day) < (birthday.month, birthday.day))
            ages.append(age)

        age_stats = {}
        if ages:
            age_counter = Counter(ages)
            age_stats = {
                "avg": round(sum(ages) / len(ages), 1),
                "min": min(ages),
                "max": max(ages),
                "buckets": [{"label": str(age), "count": count} for age, count in sorted(age_counter.items())],
            }

        return Response(
            {
                "total": total,
                "gender": {
                    "male": gender_counts.get("male", 0),
                    "female": gender_counts.get("female", 0),
                    "diverse": gender_counts.get("diverse", 0),
                    "unknown": gender_counts.get("", 0),
                },
                "by_status": by_status,
                "by_group": by_group,
                "age": age_stats,
                "can_swim": qs.filter(canSwimm=True).count(),
            }
        )

    @extend_schema(summary="Get member's parents")
    @action(detail=True, methods=["get"])
    def parents(self, request, pk=None):
        member = self.get_object()
        serializer = ParentSerializer(member.parent_set.all(), many=True, context={"request": request})
        return Response(serializer.data)

    @extend_schema(summary="Get member's events")
    @action(detail=True, methods=["get"])
    def events(self, request, pk=None):
        member = self.get_object()
        events = Event.objects.filter(member=member).order_by("-datetime")
        serializer = EventSerializer(events, many=True, context={"request": request})
        return Response(serializer.data)

    @extend_schema(summary="Get member's attachments")
    @action(detail=True, methods=["get"])
    def attachments(self, request, pk=None):
        from django.contrib.contenttypes.models import ContentType

        member = self.get_object()
        content_type = ContentType.objects.get_for_model(Member)
        attachments = Attachment.objects.filter(content_type=content_type, object_id=member.id).order_by("-uploaded_at")
        serializer = AttachmentSerializer(attachments, many=True, context={"request": request})
        return Response(serializer.data)

    @extend_schema(
        summary="Export members to Excel with column selection",
        parameters=[
            OpenApiParameter(
                "columns",
                OpenApiTypes.STR,
                description=(
                    "Comma-separated list of column keys to include. "
                    f"Available: {', '.join(MEMBER_EXPORT_COLUMNS.keys())}. "
                    "Defaults to basic member fields when omitted."
                ),
            ),
        ],
    )
    @action(detail=False, methods=["get"], url_path="export-excel", renderer_classes=[PassthroughRenderer])
    def export_excel(self, request):
        if not request.user.has_perm("members.view_member"):
            return Response({"error": "Keine Berechtigung für Mitglieder-Export"}, status=403)

        # Determine which columns to export
        columns_param = request.query_params.get("columns", "")
        if columns_param:
            requested = [c.strip() for c in columns_param.split(",") if c.strip()]
            selected_columns = [c for c in requested if c in MEMBER_EXPORT_COLUMNS]
        else:
            selected_columns = MEMBER_EXPORT_DEFAULT_COLUMNS
        if not selected_columns:
            selected_columns = MEMBER_EXPORT_DEFAULT_COLUMNS

        # Apply the same filters as the list view
        qs = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mitglieder"

        header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, col_key in enumerate(selected_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=MEMBER_EXPORT_COLUMNS[col_key])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        today = date.today()

        for row_idx, member in enumerate(qs, start=2):
            parents = list(member.parent_set.all()[:2])
            p1 = parents[0] if len(parents) > 0 else None
            p2 = parents[1] if len(parents) > 1 else None

            for col_idx, col_key in enumerate(selected_columns, start=1):
                value: str | int = ""
                if col_key == "name":
                    value = member.name
                elif col_key == "lastname":
                    value = member.lastname
                elif col_key == "gender":
                    value = {"male": "Männlich", "female": "Weiblich", "diverse": "Divers"}.get(member.gender, "")
                elif col_key == "birthday":
                    value = member.birthday.strftime("%d.%m.%Y") if member.birthday else ""
                elif col_key == "age":
                    if member.birthday:
                        value = (
                            today.year
                            - member.birthday.year
                            - ((today.month, today.day) < (member.birthday.month, member.birthday.day))
                        )
                elif col_key == "email":
                    value = member.email
                elif col_key == "phone":
                    value = member.phone
                elif col_key == "mobile":
                    value = member.mobile
                elif col_key == "street":
                    value = member.street
                elif col_key == "zip_code":
                    value = member.zip_code
                elif col_key == "city":
                    value = member.city
                elif col_key == "joined":
                    value = member.joined.strftime("%d.%m.%Y") if member.joined else ""
                elif col_key == "status":
                    value = member.status.name if member.status else ""
                elif col_key == "group":
                    value = member.group.name if member.group else ""
                elif col_key == "identityCardNumber":
                    value = member.identityCardNumber
                elif col_key == "canSwimm":
                    value = "Ja" if member.canSwimm else "Nein"
                elif col_key == "notes":
                    value = member.notes
                elif col_key == "departments":
                    value = ", ".join(d.name for d in member.departments.all())
                elif col_key == "parent1_name":
                    value = p1.name if p1 else ""
                elif col_key == "parent1_lastname":
                    value = p1.lastname if p1 else ""
                elif col_key == "parent1_email":
                    value = p1.email if p1 else ""
                elif col_key == "parent1_email2":
                    value = p1.email2 if p1 else ""
                elif col_key == "parent1_phone":
                    value = p1.phone if p1 else ""
                elif col_key == "parent1_mobile":
                    value = p1.mobile if p1 else ""
                elif col_key == "parent1_street":
                    value = p1.street if p1 else ""
                elif col_key == "parent1_zip_code":
                    value = p1.zip_code if p1 else ""
                elif col_key == "parent1_city":
                    value = p1.city if p1 else ""
                elif col_key == "parent2_name":
                    value = p2.name if p2 else ""
                elif col_key == "parent2_lastname":
                    value = p2.lastname if p2 else ""
                elif col_key == "parent2_email":
                    value = p2.email if p2 else ""
                elif col_key == "parent2_email2":
                    value = p2.email2 if p2 else ""
                elif col_key == "parent2_phone":
                    value = p2.phone if p2 else ""
                elif col_key == "parent2_mobile":
                    value = p2.mobile if p2 else ""
                elif col_key == "parent2_street":
                    value = p2.street if p2 else ""
                elif col_key == "parent2_zip_code":
                    value = p2.zip_code if p2 else ""
                elif col_key == "parent2_city":
                    value = p2.city if p2 else ""

                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (capped at 50)
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        date_str = today.strftime("%Y-%m-%d")
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="mitglieder_{date_str}.xlsx"'
        return response
