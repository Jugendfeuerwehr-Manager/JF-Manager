from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, FormView
from django.urls import reverse_lazy
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import JsonResponse, HttpResponseRedirect
from django.utils import timezone
from django.core.paginator import Paginator
from django_tables2 import RequestConfig
from django_filters.views import FilterView
from rest_framework import viewsets
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.authentication import SessionAuthentication, TokenAuthentication
from datetime import datetime, timedelta
import csv

from members.models import Member
from .models import Order, OrderableItem, OrderStatus, OrderItem, OrderItemStatusHistory
from .serializers import OrderSerializer, OrderableItemSerializer, OrderStatusSerializer, OrderItemSerializer
from .selectors import get_order_list, get_orderable_item_list, get_order_status_list
from .forms import (
    OrderForm, OrderItemFormSet, OrderFilter, OrderFilterFormHelper,
    OrderStatusUpdateForm, QuickOrderForm, BulkStatusUpdateForm, OrderItemFilterForm
)
from .tables import OrderTable
from .notifications import OrderNotificationService, OrderWorkflowService


class OrderListView(LoginRequiredMixin, FilterView):
    """Listenansicht für Bestellungen"""
    model = Order
    table_class = OrderTable
    template_name = 'orders/order_list.html'
    filterset_class = OrderFilter
    context_object_name = 'orders'
    paginate_by = 25

    def get_queryset(self):
        return get_order_list()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'].form.helper = OrderFilterFormHelper()
        
        # Table für django-tables2
        table = OrderTable(self.filterset.qs)
        RequestConfig(self.request).configure(table)
        context['table'] = table
        
        # Statistiken
        queryset = self.filterset.qs
        context['stats'] = {
            'total_orders': queryset.count(),
            'pending_items': OrderItem.objects.filter(
                order__in=queryset,
                status__code__in=['NEW', 'ORDERED']
            ).count(),
            'delivered_items': OrderItem.objects.filter(
                order__in=queryset,
                status__code='DELIVERED'
            ).count(),
        }
        
        return context


class OrderDetailView(LoginRequiredMixin, DetailView):
    """Detailansicht für Bestellungen"""
    model = Order
    template_name = 'orders/order_detail.html'
    context_object_name = 'order'

    def get_queryset(self):
        return Order.objects.select_related('member', 'ordered_by').prefetch_related(
            'items__item', 'items__status'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_manage_orders'] = self.request.user.has_perm('orders.can_manage_orders')
        context['can_change_status'] = self.request.user.has_perm('orders.can_change_order_status')
        context['order'].next_status_options = self.object.get_next_status_options()
        return context


class OrderUpdateView(LoginRequiredMixin, UpdateView):
    """Ansicht zum Bearbeiten von Bestellungen"""
    model = Order
    form_class = OrderForm
    template_name = 'orders/order_update.html'
    context_object_name = 'order'

    def get_success_url(self):
        return reverse_lazy('orders:detail', kwargs={'pk': self.object.pk})

    def get_queryset(self):
        return Order.objects.select_related('member', 'ordered_by').prefetch_related(
            'items__item', 'items__status'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = OrderItemFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = OrderItemFormSet(instance=self.object)
        context['orderable_items'] = OrderableItem.objects.filter(is_active=True)
        context['can_manage_orders'] = self.request.user.has_perm('orders.can_manage_orders')
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        with transaction.atomic():
            self.object = form.save()
            
            if formset.is_valid():
                formset.instance = self.object
                items = formset.save(commit=False)
                
                # Default Status für neue Items
                default_status = OrderStatus.objects.filter(code='NEW').first()
                if not default_status:
                    default_status = OrderStatus.objects.filter(code='ORDERED').first()
                if not default_status:
                    default_status = OrderStatus.objects.first()
                
                for item in items:
                    if not item.status_id:
                        item.status = default_status
                    item.save()
                
                # Gelöschte Items entfernen
                for item in formset.deleted_objects:
                    item.delete()
                
                formset.save_m2m()
                
                messages.success(
                    self.request, 
                    f'Bestellung für {self.object.member} wurde erfolgreich aktualisiert.'
                )
                return HttpResponseRedirect(self.get_success_url())
            else:
                return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Bitte korrigieren Sie die Fehler im Formular.')
        return super().form_invalid(form)


class OrderCreateView(LoginRequiredMixin, CreateView):
    """Ansicht zum Erstellen neuer Bestellungen"""
    model = Order
    form_class = OrderForm
    template_name = 'orders/order_create.html'
    success_url = reverse_lazy('orders:list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = OrderItemFormSet(self.request.POST)
        else:
            context['formset'] = OrderItemFormSet()
        context['orderable_items'] = OrderableItem.objects.filter(is_active=True)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        with transaction.atomic():
            form.instance.ordered_by = self.request.user
            self.object = form.save()
            
            if formset.is_valid():
                formset.instance = self.object
                items = formset.save(commit=False)
                
                # Default Status für neue Items setzen - beginne mit "Neu"
                default_status = OrderStatus.objects.filter(code='NEW').first()
                if not default_status:
                    # Fallback auf ORDERED falls NEW noch nicht existiert
                    default_status = OrderStatus.objects.filter(code='ORDERED').first()
                if not default_status:
                    default_status = OrderStatus.objects.first()
                
                for item in items:
                    if not item.status_id:
                        item.status = default_status
                    item.save()
                
                formset.save_m2m()
                
                # Send notification for new order
                OrderNotificationService.send_order_created_notification(self.object, self.request)
                
                messages.success(
                    self.request, 
                    f'Bestellung für {self.object.member} wurde erfolgreich erstellt.'
                )
                return HttpResponseRedirect(self.get_success_url())
            else:
                return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Bitte korrigieren Sie die Fehler im Formular.')
        return super().form_invalid(form)


class QuickOrderCreateView(LoginRequiredMixin, FormView):
    """Schnelle Bestellung für häufig bestellte Artikel"""
    form_class = QuickOrderForm
    template_name = 'orders/quick_order.html'
    success_url = reverse_lazy('orders:list')

    def form_valid(self, form):
        member = form.cleaned_data['member']
        
        with transaction.atomic():
            # Neue Bestellung erstellen
            order = Order.objects.create(
                member=member,
                ordered_by=self.request.user,
                notes='Schnellbestellung'
            )
            
            # Default Status für neue Items - beginne mit "Neu"
            default_status = OrderStatus.objects.filter(code='NEW').first()
            if not default_status:
                # Fallback auf ORDERED falls NEW noch nicht existiert
                default_status = OrderStatus.objects.filter(code='ORDERED').first()
            if not default_status:
                default_status = OrderStatus.objects.first()
            
            # Ausgewählte Items hinzufügen
            items_added = 0
            for field_name, value in form.cleaned_data.items():
                if field_name.startswith('item_') and value:
                    item_id = field_name.split('_')[1]
                    item = OrderableItem.objects.get(id=item_id)
                    
                    # Größe prüfen
                    size_field = f'size_{item_id}'
                    size = form.cleaned_data.get(size_field, '')
                    
                    OrderItem.objects.create(
                        order=order,
                        item=item,
                        size=size,
                        quantity=1,
                        status=default_status
                    )
                    items_added += 1
            
            if items_added > 0:
                # Send notification for new quick order
                OrderNotificationService.send_order_created_notification(order, self.request)
                
                messages.success(
                    self.request, 
                    f'Schnellbestellung für {member} mit {items_added} Artikel(n) wurde erstellt.'
                )
            else:
                order.delete()
                messages.warning(self.request, 'Keine Artikel ausgewählt.')
                return self.form_invalid(form)
        
        return HttpResponseRedirect(self.get_success_url())


@login_required
@permission_required('orders.can_change_order_status', raise_exception=True)
def update_order_item_status(request, order_id, item_id):
    """Status eines Bestellartikels aktualisieren"""
    order = get_object_or_404(Order, id=order_id)
    order_item = get_object_or_404(OrderItem, id=item_id, order=order)
    
    if request.method == 'POST':
        form = OrderStatusUpdateForm(request.POST, instance=order_item)
        if form.is_valid():
            old_status = order_item.status
            # Save with tracking information
            updated_item = form.save(commit=False)
            updated_item.save(changed_by=request.user, status_change_notes=f"Status updated via web interface by {request.user.get_full_name()}")
            
            # Send notification if status changed
            if old_status != updated_item.status:
                OrderNotificationService.send_status_update_notification(
                    updated_item, old_status, updated_item.status, request.user, request
                )
            
            messages.success(request, f'Status für {order_item.item.name} wurde aktualisiert.')
            return redirect('orders:detail', pk=order.pk)
    else:
        form = OrderStatusUpdateForm(instance=order_item)
    
    return render(request, 'orders/update_status.html', {
        'form': form,
        'order': order,
        'order_item': order_item,
    })


@login_required
def get_item_sizes(request, item_id):
    """AJAX Endpoint für Größen eines Artikels"""
    try:
        item = OrderableItem.objects.get(id=item_id, is_active=True)
        sizes = item.get_sizes_list() if item.has_sizes else []
        return JsonResponse({
            'has_sizes': item.has_sizes,
            'sizes': sizes
        })
    except OrderableItem.DoesNotExist:
        return JsonResponse({'error': 'Artikel nicht gefunden'}, status=404)


############################ API ################################################

class OrderViewSet(viewsets.ModelViewSet):
    """
    API Endpoint für Bestellungen
    """
    authentication_classes = [JWTAuthentication, TokenAuthentication, SessionAuthentication]
    queryset = get_order_list()
    serializer_class = OrderSerializer
    filterset_fields = ['member', 'ordered_by', 'order_date']
    search_fields = ['member__name', 'member__lastname', 'notes']


class OrderableItemViewSet(viewsets.ModelViewSet):
    """
    API Endpoint für bestellbare Artikel
    """
    authentication_classes = [JWTAuthentication, TokenAuthentication, SessionAuthentication]
    queryset = get_orderable_item_list()
    serializer_class = OrderableItemSerializer
    filterset_fields = ['category', 'has_sizes', 'is_active']
    search_fields = ['name', 'category', 'description']


class OrderStatusViewSet(viewsets.ModelViewSet):
    """
    API Endpoint für Bestellstatus
    """
    authentication_classes = [JWTAuthentication, TokenAuthentication, SessionAuthentication]
    queryset = get_order_status_list()
    serializer_class = OrderStatusSerializer
    filterset_fields = ['is_active']
    search_fields = ['name', 'code']


class OrderItemViewSet(viewsets.ModelViewSet):
    """
    API Endpoint für Bestellartikel
    """
    authentication_classes = [JWTAuthentication, TokenAuthentication, SessionAuthentication]
    queryset = OrderItem.objects.select_related('order', 'item', 'status')
    serializer_class = OrderItemSerializer
    filterset_fields = ['order', 'item', 'status', 'size']
    search_fields = ['item__name', 'notes']


############################ BULK OPERATIONS ####################################

@login_required
@permission_required('orders.can_change_order_status', raise_exception=True)
@login_required
@permission_required('orders.can_manage_orders', raise_exception=True)
def quick_order_status_change(request, pk):
    """Schneller Statuswechsel für eine Bestellung"""
    order = get_object_or_404(Order, pk=pk)
    new_status_id = request.GET.get('status')
    
    if not new_status_id:
        messages.error(request, 'Kein Status ausgewählt.')
        return redirect('orders:detail', pk=pk)
    
    new_status = get_object_or_404(OrderStatus, pk=new_status_id)
    
    # Alle Artikel der Bestellung auf den neuen Status setzen
    updated_count = 0
    with transaction.atomic():
        for item in order.items.all():
            item.status = new_status
            item.save(changed_by=request.user, status_change_notes=f"Schnellwechsel zu {new_status.name}")
            updated_count += 1
    
    messages.success(
        request, 
        f'Status aller {updated_count} Artikel wurde auf "{new_status.name}" geändert.'
    )
    return redirect('orders:detail', pk=pk)


def bulk_status_update(request):
    """View for bulk status updates of order items"""
    filter_form = OrderItemFilterForm(request.GET or None)
    
    # Get initial queryset
    queryset = OrderItem.objects.select_related('order', 'item', 'status', 'order__member')
    
    # Apply filters
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('status'):
            queryset = queryset.filter(status=filter_form.cleaned_data['status'])
        if filter_form.cleaned_data.get('item_category'):
            queryset = queryset.filter(item__category=filter_form.cleaned_data['item_category'])
        if filter_form.cleaned_data.get('member'):
            queryset = queryset.filter(order__member=filter_form.cleaned_data['member'])
        if filter_form.cleaned_data.get('date_from'):
            queryset = queryset.filter(order__order_date__date__gte=filter_form.cleaned_data['date_from'])
        if filter_form.cleaned_data.get('date_to'):
            queryset = queryset.filter(order__order_date__date__lte=filter_form.cleaned_data['date_to'])
    
    bulk_form = None
    if request.method == 'POST':
        bulk_form = BulkStatusUpdateForm(request.POST, queryset=queryset)
        if bulk_form.is_valid():
            order_items = bulk_form.cleaned_data['order_items']
            new_status = bulk_form.cleaned_data['new_status']
            update_dates = bulk_form.cleaned_data['update_dates']
            notes = bulk_form.cleaned_data['notes']
            
            updated_count = 0
            with transaction.atomic():
                for item in order_items:
                    old_status = item.status
                    item.status = new_status
                    
                    # Update dates if requested
                    if update_dates:
                        if new_status.code in ['received', 'eingegangen']:
                            from django.utils import timezone
                            item.received_date = timezone.now()
                        elif new_status.code in ['delivered', 'ausgegeben']:
                            from django.utils import timezone
                            item.delivered_date = timezone.now()
                    
                    # Add notes if provided
                    if notes:
                        if item.notes:
                            item.notes += f"\n{notes}"
                        else:
                            item.notes = notes
                    
                    # Save with tracking information
                    bulk_notes = f"Bulk status update by {request.user.get_full_name()}"
                    if notes:
                        bulk_notes += f": {notes}"
                    item.save(changed_by=request.user, status_change_notes=bulk_notes)
                    updated_count += 1
                
                # Send bulk notification
                OrderNotificationService.send_bulk_status_update_notification(
                    order_items, new_status, request.user, request
                )
            
            messages.success(
                request, 
                f'{updated_count} Artikel wurden erfolgreich auf Status "{new_status.name}" aktualisiert.'
            )
            return redirect('orders:bulk_status_update')
    else:
        bulk_form = BulkStatusUpdateForm(queryset=queryset)
    
    return render(request, 'orders/bulk_status_update.html', {
        'filter_form': filter_form,
        'bulk_form': bulk_form,
        'queryset': queryset[:100],  # Limit display for performance
        'queryset_count': queryset.count(),
    })


@login_required
@permission_required('orders.can_manage_orders', raise_exception=True)
def order_analytics_dashboard(request):
    """Dashboard with order analytics and reports"""
    from django.db.models import Count, Q, Avg
    from django.utils import timezone
    from datetime import timedelta
    
    # Date ranges
    today = timezone.now().date()
    last_30_days = today - timedelta(days=30)
    last_90_days = today - timedelta(days=90)
    
    # Basic statistics
    stats = {
        'total_orders': Order.objects.count(),
        'orders_last_30_days': Order.objects.filter(order_date__date__gte=last_30_days).count(),
        'total_items': OrderItem.objects.count(),
        'pending_items': OrderItem.objects.filter(status__code='pending').count(),
        'delivered_items': OrderItem.objects.filter(status__code='delivered').count(),
    }
    
    # Status distribution
    status_data = (OrderItem.objects
                  .values('status__name', 'status__color')
                  .annotate(count=Count('id'))
                  .order_by('-count'))
    
    # Category distribution
    category_data = (OrderItem.objects
                    .values('item__category')
                    .annotate(count=Count('id'))
                    .order_by('-count'))
    
    # Recent orders
    recent_orders = (Order.objects
                    .select_related('member', 'ordered_by')
                    .prefetch_related('items__status')
                    .order_by('-order_date')[:10])
    
    # Monthly order trends (last 12 months)
    monthly_data = []
    for i in range(12):
        month_start = (today.replace(day=1) - timedelta(days=30*i)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        count = Order.objects.filter(
            order_date__date__gte=month_start,
            order_date__date__lte=month_end
        ).count()
        
        monthly_data.append({
            'month': month_start.strftime('%Y-%m'),
            'count': count
        })
    
    monthly_data.reverse()
    
    return render(request, 'orders/analytics_dashboard.html', {
        'stats': stats,
        'status_data': status_data,
        'category_data': category_data,
        'recent_orders': recent_orders,
        'monthly_data': monthly_data,
    })


@login_required
def export_orders(request):
    """Export orders to CSV/Excel"""
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    
    # Get export parameters
    format_type = request.GET.get('format', 'csv')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status_filter = request.GET.get('status')
    
    # Build queryset
    queryset = Order.objects.select_related('member', 'ordered_by').prefetch_related('items__status', 'items__item')
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(order_date__date__gte=date_from)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            queryset = queryset.filter(order_date__date__lte=date_to)
        except ValueError:
            pass
    
    if status_filter:
        queryset = queryset.filter(items__status__id=status_filter).distinct()
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'bestellungen_{timestamp}.{format_type}'
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Bestellung Nr.', 'Mitglied', 'Bestelldatum', 'Bestellt von',
            'Artikel', 'Kategorie', 'Größe', 'Anzahl', 'Status',
            'Eingangsdatum', 'Ausgabedatum', 'Bemerkungen'
        ])
        
        for order in queryset:
            for item in order.items.all():
                writer.writerow([
                    order.pk,
                    order.member.get_full_name(),
                    order.order_date.strftime('%d.%m.%Y %H:%M'),
                    order.ordered_by.get_full_name() if order.ordered_by else '',
                    item.item.name,
                    item.item.category,
                    item.size or '',
                    item.quantity,
                    item.status.name,
                    item.received_date.strftime('%d.%m.%Y %H:%M') if item.received_date else '',
                    item.delivered_date.strftime('%d.%m.%Y %H:%M') if item.delivered_date else '',
                    item.notes or ''
                ])
        
        return response
    
    else:  # Excel format
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Bestellungen"
            
            # Header row
            headers = [
                'Bestellung Nr.', 'Mitglied', 'Bestelldatum', 'Bestellt von',
                'Artikel', 'Kategorie', 'Größe', 'Anzahl', 'Status',
                'Eingangsdatum', 'Ausgabedatum', 'Bemerkungen'
            ]
            
            # Style header
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data rows
            row = 2
            for order in queryset:
                for item in order.items.all():
                    ws.cell(row=row, column=1, value=order.pk)
                    ws.cell(row=row, column=2, value=order.member.get_full_name())
                    ws.cell(row=row, column=3, value=order.order_date.strftime('%d.%m.%Y %H:%M'))
                    ws.cell(row=row, column=4, value=order.ordered_by.get_full_name() if order.ordered_by else '')
                    ws.cell(row=row, column=5, value=item.item.name)
                    ws.cell(row=row, column=6, value=item.item.category)
                    ws.cell(row=row, column=7, value=item.size or '')
                    ws.cell(row=row, column=8, value=item.quantity)
                    ws.cell(row=row, column=9, value=item.status.name)
                    ws.cell(row=row, column=10, value=item.received_date.strftime('%d.%m.%Y %H:%M') if item.received_date else '')
                    ws.cell(row=row, column=11, value=item.delivered_date.strftime('%d.%m.%Y %H:%M') if item.delivered_date else '')
                    ws.cell(row=row, column=12, value=item.notes or '')
                    row += 1
            
            # Auto-adjust column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except ImportError:
            messages.error(request, 'Excel-Export nicht verfügbar. Bitte installieren Sie openpyxl.')
            return redirect('orders:list')


@login_required
def bulk_status_update_view(request):
    """View for bulk status updates"""
    filter_form = OrderItemFilterForm(request.GET or None)
    
    # Get initial queryset
    queryset = OrderItem.objects.select_related('order', 'item', 'status', 'order__member')
    
    # Apply filters
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('status'):
            queryset = queryset.filter(status=filter_form.cleaned_data['status'])
        if filter_form.cleaned_data.get('item_category'):
            queryset = queryset.filter(item__category=filter_form.cleaned_data['item_category'])
        if filter_form.cleaned_data.get('member'):
            queryset = queryset.filter(order__member=filter_form.cleaned_data['member'])
        if filter_form.cleaned_data.get('date_from'):
            queryset = queryset.filter(order__order_date__date__gte=filter_form.cleaned_data['date_from'])
        if filter_form.cleaned_data.get('date_to'):
            queryset = queryset.filter(order__order_date__date__lte=filter_form.cleaned_data['date_to'])
    
    if request.method == 'POST':
        form = BulkStatusUpdateForm(request.POST, queryset=queryset)
        if form.is_valid():
            order_items = form.cleaned_data['order_items']
            new_status = form.cleaned_data['new_status']
            update_dates = form.cleaned_data['update_dates']
            notes = form.cleaned_data['notes']
            
            updated_count = 0
            with transaction.atomic():
                for item in order_items:
                    old_status = item.status
                    item.status = new_status
                    
                    # Update dates if requested
                    if update_dates:
                        if new_status.code in ['received', 'eingegangen']:
                            item.received_date = timezone.now()
                        elif new_status.code in ['delivered', 'ausgegeben']:
                            item.delivered_date = timezone.now()
                    
                    # Add notes if provided
                    if notes:
                        if item.notes:
                            item.notes += f"\n{notes}"
                        else:
                            item.notes = notes
                    
                    # Save with tracking information
                    bulk_notes = f"Bulk status update by {request.user.get_full_name()}"
                    if notes:
                        bulk_notes += f": {notes}"
                    item.save(changed_by=request.user, status_change_notes=bulk_notes)
                    updated_count += 1
                
                # Send bulk notification
                OrderNotificationService.send_bulk_status_update_notification(
                    order_items, new_status, request.user, request
                )
            
            messages.success(
                request, 
                f'{updated_count} Artikel wurden erfolgreich auf Status "{new_status.name}" aktualisiert.'
            )
            return redirect('orders:bulk_status_update')
    else:
        form = BulkStatusUpdateForm(queryset=queryset)
    
    return render(request, 'orders/bulk_status_update.html', {
        'filter_form': filter_form,
        'bulk_form': form,
        'queryset': queryset[:100],  # Limit display for performance
        'queryset_count': queryset.count(),
    })


@login_required 
def analytics_dashboard_view(request):
    """Advanced analytics dashboard for orders"""
    try:
        import pandas as pd
        PANDAS_AVAILABLE = True
    except ImportError:
        PANDAS_AVAILABLE = False
    
    import json
    
    if PANDAS_AVAILABLE:
        # Date range for analysis
        end_date = timezone.now()
        start_date = end_date - timedelta(days=90)  # Last 90 days
        
        # Basic statistics
        total_orders = Order.objects.count()
        recent_orders = Order.objects.filter(order_date__gte=start_date).count()
        
        # Status distribution
        status_distribution = list(
            OrderItem.objects.values('status__name', 'status__color')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        
        # Orders over time (last 12 months)
        orders_over_time = []
        for i in range(12):
            month_start = (end_date - timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            count = Order.objects.filter(
                order_date__gte=month_start,
                order_date__lte=month_end
            ).count()
            orders_over_time.append({
                'month': month_start.strftime('%Y-%m'),
                'count': count
            })
        orders_over_time.reverse()
        
        # Category distribution
        category_distribution = list(
            OrderItem.objects.values('item__category')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        
        # Member activity (top 10)
        member_activity = list(
            Order.objects.values('member__name', 'member__lastname')
            .annotate(order_count=Count('id'), item_count=Sum('items__quantity'))
            .order_by('-order_count')[:10]
        )
        
        # Processing time analysis (for completed items)
        completed_items = OrderItem.objects.filter(
            delivered_date__isnull=False,
            order__order_date__gte=start_date
        ).annotate(
            processing_days=Count('id')  # Simplified - would need custom calculation
        )
        
        avg_processing_time = None
        if completed_items.exists():
            avg_processing_time = 7  # Placeholder - would need proper calculation
        
        # Pending items by age
        pending_statuses = OrderStatus.objects.filter(code__in=['NEW', 'pending', 'ordered'])
        pending_items = OrderItem.objects.filter(status__in=pending_statuses)
        
        pending_by_age = {
            'under_week': pending_items.filter(order__order_date__gte=end_date - timedelta(days=7)).count(),
            'week_to_month': pending_items.filter(
                order__order_date__gte=end_date - timedelta(days=30),
                order__order_date__lt=end_date - timedelta(days=7)
            ).count(),
            'over_month': pending_items.filter(order__order_date__lt=end_date - timedelta(days=30)).count(),
        }
        
        context = {
            'total_orders': total_orders,
            'recent_orders': recent_orders,
            'status_distribution': json.dumps(status_distribution),
            'orders_over_time': json.dumps(orders_over_time),
            'category_distribution': json.dumps(category_distribution),
            'member_activity': member_activity,
            'avg_processing_time': round(avg_processing_time, 1) if avg_processing_time else None,
            'pending_by_age': pending_by_age,
            'total_pending': sum(pending_by_age.values()),
        }
        
    else:
        # Pandas not available - provide basic statistics only
        context = {
            'pandas_missing': True,
            'total_orders': Order.objects.count(),
            'recent_orders': Order.objects.filter(
                order_date__gte=timezone.now() - timedelta(days=30)
            ).count(),
        }
    
    return render(request, 'orders/analytics_dashboard.html', context)


@login_required
def export_orders_csv(request):
    """Export orders to CSV format"""
    import csv
    from django.http import HttpResponse
    
    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="orders_export.csv"'
    
    writer = csv.writer(response)
    
    # Write header row
    writer.writerow([
        'Bestellung ID', 'Mitglied', 'Bestelldatum', 'Bestellt von',
        'Artikel', 'Kategorie', 'Größe', 'Anzahl', 'Status',
        'Eingangsdatum', 'Ausgabedatum', 'Bemerkungen'
    ])
    
    # Get filter parameters
    filter_form = OrderItemFilterForm(request.GET)
    items = OrderItem.objects.select_related('order', 'item', 'status', 'order__member')
    
    if filter_form.is_valid():
        if filter_form.cleaned_data['status']:
            items = items.filter(status=filter_form.cleaned_data['status'])
        if filter_form.cleaned_data['item_category']:
            items = items.filter(item__category=filter_form.cleaned_data['item_category'])
        if filter_form.cleaned_data['member']:
            items = items.filter(order__member=filter_form.cleaned_data['member'])
        if filter_form.cleaned_data['date_from']:
            items = items.filter(order__order_date__gte=filter_form.cleaned_data['date_from'])
        if filter_form.cleaned_data['date_to']:
            items = items.filter(order__order_date__lte=filter_form.cleaned_data['date_to'])
    
    # Write data rows
    for item in items:
        writer.writerow([
            item.order.pk,
            item.order.member.get_full_name(),
            item.order.order_date.strftime('%d.%m.%Y %H:%M'),
            item.order.ordered_by.get_full_name() if item.order.ordered_by else '',
            item.item.name,
            item.item.category,
            item.size or '',
            item.quantity,
            item.status.name,
            item.received_date.strftime('%d.%m.%Y %H:%M') if item.received_date else '',
            item.delivered_date.strftime('%d.%m.%Y %H:%M') if item.delivered_date else '',
            item.notes or ''
        ])
    
    return response


@login_required
def order_items_list_view(request):
    """Enhanced list view for order items with filtering and bulk operations"""
    filter_form = OrderItemFilterForm(request.GET)
    items = OrderItem.objects.select_related('order', 'item', 'status', 'order__member').order_by('-order__order_date')
    
    if filter_form.is_valid():
        if filter_form.cleaned_data['status']:
            items = items.filter(status=filter_form.cleaned_data['status'])
        if filter_form.cleaned_data['item_category']:
            items = items.filter(item__category=filter_form.cleaned_data['item_category'])
        if filter_form.cleaned_data['member']:
            items = items.filter(order__member=filter_form.cleaned_data['member'])
        if filter_form.cleaned_data['date_from']:
            items = items.filter(order__order_date__gte=filter_form.cleaned_data['date_from'])
        if filter_form.cleaned_data['date_to']:
            items = items.filter(order__order_date__lte=filter_form.cleaned_data['date_to'])
    
    # Pagination
    paginator = Paginator(items, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics for filtered items
    stats = {
        'total_items': items.count(),
        'by_status': items.values('status__name').annotate(count=Count('id')),
        'by_category': items.values('item__category').annotate(count=Count('id')),
    }
    
    context = {
        'filter_form': filter_form,
        'page_obj': page_obj,
        'stats': stats,
        'bulk_form': BulkStatusUpdateForm(queryset=items),
    }
    
    return render(request, 'orders/order_items_list.html', context)
